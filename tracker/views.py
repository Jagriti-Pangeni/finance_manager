import re
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import UserProfile, Transaction, Category, Account
from django.utils import timezone
from django.contrib.auth.models import User
from django.contrib.auth import login
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation
from django.contrib.auth import logout
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from django.db.models import Sum

def home(request):
    return render(request, "tracker/homepage.html")


def loginpage(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, "Invalid username or password")
            # ✅ render instead of redirect — keeps username in form
            return render(request, "tracker/login.html", {
                "username": username
            })

    return render(request, "tracker/login.html")


def register(request):
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        email = request.POST.get("eMail", "").strip().lower()
        password = request.POST.get("password", "")
        confirm_password = request.POST.get("conf_password", "")

        # Bundle form data to re-fill fields on error
        form_data = {
            "name": name,
            "eMail": email,
        }

        def error(msg):
            messages.error(request, msg)
            return render(request, "tracker/register.html", form_data)

        # ── USERNAME VALIDATION ──────────────────────────────────────────
        if not name:
            return error("Username is required.")

        if len(name) < 3 or len(name) > 20:
            return error("Username must be between 3 and 20 characters.")

        if not re.match(r'^[a-zA-Z0-9_ .-]+$', name):
            return error("Username can only contain letters, numbers, underscores (_), dots (.), and hyphens (-).")

        if re.match(r'^[_.\-]', name) or re.match(r'[_.\-]$', name):
            return error("Username cannot start or end with a special character.")

        if re.search(r'[_.\-]{2,}', name):
            return error("Username cannot have consecutive special characters.")
        
        if not name.replace(" ", ""):
            return error("Username cannot be only spaces.")

        RESERVED_USERNAMES = ['admin', 'root', 'superuser', 'administrator', 'support', 'help', 'null', 'undefined']
        if name.lower() in RESERVED_USERNAMES:
            return error("This username is reserved. Please choose another.")

        if User.objects.filter(username__iexact=name).exists():
            return error("Username already exists.")

        # ── EMAIL VALIDATION ─────────────────────────────────────────────
        if not email:
            return error("Email is required.")

        email_regex = r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$'
        if not re.match(email_regex, email):
            return error("Enter a valid email address.")

        BLOCKED_EMAIL_DOMAINS = ['tempmail.com', 'throwaway.email', 'mailinator.com', 'guerrillamail.com']
        email_domain = email.split('@')[-1]
        if email_domain in BLOCKED_EMAIL_DOMAINS:
            return error("Disposable email addresses are not allowed.")

        if User.objects.filter(email__iexact=email).exists():
            return error("An account with this email already exists.")

        # ── PASSWORD VALIDATION ──────────────────────────────────────────
        if not password:
            return error("Password is required.")
        
        if ' ' in password:
            return error("Password cannot contain spaces.")

        if len(password) < 8:
            return error("Password must be at least 8 characters long.")

        if len(password) > 128:
            return error("Password must not exceed 128 characters.")

        if not re.search(r'\d', password):
            return error("Password must contain at least one number.")

        COMMON_PASSWORDS = ['password', '12345678', 'password1', 'qwerty123', 'iloveyou']
        if password.lower() in COMMON_PASSWORDS:
            return error("This password is too common. Please choose a stronger one.")

        if name.lower() in password.lower():
            return error("Password should not contain your username.")

        if password != confirm_password:
            return error("Passwords do not match.")

        # ── CREATE USER ──────────────────────────────────────────────────
        user = User.objects.create_user(username=name, email=email, password=password)
        user.save()

        Account.objects.create(user=user, balance=Decimal('0.00'))

        Category.objects.bulk_create([
            Category(name="Salary", type="income", user=user),
            Category(name="Freelance", type="income", user=user),
            Category(name="Food", type="expense", user=user),
            Category(name="Transport", type="expense", user=user),
            Category(name="Bills", type="expense", user=user),
        ])

        login(request, user)
        messages.success(request, "Account created successfully!")
        return redirect('dashboard')

    return render(request, "tracker/register.html")


@login_required(login_url='/login/')
def dashboard(request):
    transactions = Transaction.objects.filter(user=request.user).order_by('-date')[:10]
    total_income = sum(t.amount for t in transactions if t.transaction_type == 'income')
    total_expense = sum(t.amount for t in transactions if t.transaction_type == 'expense')
    balance = total_income - total_expense

    income_categories = Category.objects.filter(user=request.user, type='income')
    expense_categories = Category.objects.filter(user=request.user, type='expense')

    return render(request, "tracker/dashboard.html", {
        "transactions": transactions,
        "total_income": total_income,
        "total_expense": total_expense,
        "balance": balance,
        "income_categories": income_categories,
        "expense_categories": expense_categories
    })


@login_required(login_url='/login/')
def add_income(request):
    if request.method == "POST":
        category_id = request.POST.get("category")
        amount = request.POST.get("amount")
        date_str = request.POST.get("date")

        if not amount:
            messages.error(request, "Amount is required.")
            return redirect('dashboard')

        try:
            amount = Decimal(amount)
        except InvalidOperation:
            messages.error(request, "Enter a valid numeric amount.")
            return redirect('dashboard')

        if amount <= 0:
            messages.error(request, "Income amount must be greater than zero.")
            return redirect('dashboard')

        if amount > Decimal('9999999.99'):
            messages.error(request, "Income amount is unrealistically large.")
            return redirect('dashboard')

        if round(amount, 2) != amount:
            messages.error(request, "Amount can have at most 2 decimal places.")
            return redirect('dashboard')

        if not date_str:
            messages.error(request, "Date is required.")
            return redirect('dashboard')

        try:
            income_date = date.fromisoformat(date_str)
        except ValueError:
            messages.error(request, "Enter a valid date in YYYY-MM-DD format.")
            return redirect('dashboard')

        today = date.today()
        if income_date > today + timedelta(days=1):
            messages.error(request, "Income date cannot be more than 1 day in the future.")
            return redirect('dashboard')

        if income_date < today - timedelta(days=365):
            messages.error(request, "Income date cannot be more than 1 year in the past.")
            return redirect('dashboard')

        if not category_id:
            messages.error(request, "Category is required.")
            return redirect('dashboard')

        try:
            category = Category.objects.get(id=category_id)
        except Category.DoesNotExist:
            messages.error(request, "Selected category does not exist.")
            return redirect('dashboard')

        account = Account.objects.filter(user=request.user).first()
        if not account:
            messages.error(request, "No account found. Please create an account first.")
            return redirect('dashboard')

        Transaction.objects.create(
            user=request.user,
            account=account,
            category=category,
            amount=amount,
            transaction_type='income',
            date=income_date
        )
        account.balance += amount
        account.save()
        messages.success(request, f"Income of {amount} added successfully.")

    return redirect('dashboard')
@login_required(login_url='/login/')
def logout_user(request):
    logout(request)
    messages.success(request, "You have been logged out successfully.")
    return redirect("login")


@login_required(login_url='/login/')
def add_expense(request):
    if request.method == "POST":
        category_id = request.POST.get("category")
        amount = request.POST.get("amount")
        date_str = request.POST.get("date")

        if not amount:
            messages.error(request, "Amount is required.")
            return redirect('dashboard')

        try:
            amount = Decimal(amount)
        except InvalidOperation:
            messages.error(request, "Enter a valid numeric amount.")
            return redirect('dashboard')

        if amount <= 0:
            messages.error(request, "Expense amount must be greater than zero.")
            return redirect('dashboard')

        if amount > Decimal('9999999.99'):
            messages.error(request, "Expense amount is unrealistically large.")
            return redirect('dashboard')

        if round(amount, 2) != amount:
            messages.error(request, "Amount can have at most 2 decimal places.")
            return redirect('dashboard')

        if not date_str:
            messages.error(request, "Date is required.")
            return redirect('dashboard')

        try:
            expense_date = date.fromisoformat(date_str)
        except ValueError:
            messages.error(request, "Enter a valid date in YYYY-MM-DD format.")
            return redirect('dashboard')

        today = date.today()
        if expense_date > today + timedelta(days=1):
            messages.error(request, "Expense date cannot be more than 1 day in the future.")
            return redirect('dashboard')

        if expense_date < today - timedelta(days=365):
            messages.error(request, "Expense date cannot be more than 1 year in the past.")
            return redirect('dashboard')

        if not category_id:
            messages.error(request, "Category is required.")
            return redirect('dashboard')

        try:
            category = Category.objects.get(id=category_id)
        except Category.DoesNotExist:
            messages.error(request, "Selected category does not exist.")
            return redirect('dashboard')

        account = Account.objects.filter(user=request.user).first()
        if not account:
            messages.error(request, "No account found. Please create an account first.")
            return redirect('dashboard')

        if account.balance < amount:
            messages.error(request, f"Insufficient balance. Your current balance is {account.balance}.")
            return redirect('dashboard')

        Transaction.objects.create(
            user=request.user,
            account=account,
            category=category,
            amount=amount,
            transaction_type='expense',
            date=expense_date
        )
        messages.success(request, f"Expense of {amount} added successfully.")
        account.balance -= amount
        account.save()

    return redirect('dashboard')


@login_required(login_url="/login/")
def add_category(request):

    if request.method == "POST":
        name = request.POST.get("name").strip()
        category_type = request.POST.get("type")

        if Category.objects.filter(
            user=request.user,
            name__iexact=name,
            type=category_type
        ).exists():
            messages.error(request, "Category already exists.")
            return redirect("dashboard")

        Category.objects.create(
            user=request.user,
            name=name,
            type=category_type
        )

        messages.success(request, "Category added successfully.")

    return redirect("dashboard")


@login_required(login_url="/login/")
def profile(request):
    profile, created = UserProfile.objects.get_or_create(user=request.user)

    if request.method == "POST":
        request.user.username = request.POST.get("username")
        request.user.email = request.POST.get("email")
        request.user.save()

        messages.success(request, "Profile updated successfully.")
        return redirect("profile")

    return render(request, "tracker/profile.html", {
        "profile": profile
    })


@login_required(login_url='/login/')
def reports(request):
    period = request.GET.get('period', 'monthly')
    chart_type = request.GET.get('chart_type', 'bar')

    today = date.today()

    if period == 'daily':
        start_date = today
    elif period == 'weekly':
        start_date = today - timedelta(days=7)
    else:
        start_date = today - timedelta(days=30)

    transactions = Transaction.objects.filter(
        user=request.user,
        date__gte=start_date,
        date__lte=today
    )

    expense_by_category = (
        transactions.filter(transaction_type='expense')
        .values('category__name')
        .annotate(total=Sum('amount'))
        .order_by('-total')
    )

    income_by_category = (
        transactions.filter(transaction_type='income')
        .values('category__name')
        .annotate(total=Sum('amount'))
        .order_by('-total')
    )

    total_income = transactions.filter(transaction_type='income').aggregate(Sum('amount'))['amount__sum'] or 0
    total_expense = transactions.filter(transaction_type='expense').aggregate(Sum('amount'))['amount__sum'] or 0
    net_balance = total_income - total_expense

    expense_labels = [item['category__name'] for item in expense_by_category]
    expense_data = [float(item['total']) for item in expense_by_category]
    income_labels = [item['category__name'] for item in income_by_category]
    income_data = [float(item['total']) for item in income_by_category]

    return render(request, 'tracker/reports.html', {
        'period': period,
        'chart_type': chart_type,
        'total_income': total_income,
        'total_expense': total_expense,
        'net_balance': net_balance,
        'expense_labels': expense_labels,
        'expense_data': expense_data,
        'income_labels': income_labels,
        'income_data': income_data,
        'transactions': transactions.order_by('-date'),
    })


@login_required(login_url='/login/')
def export_csv(request):
    period = request.GET.get('period', 'monthly')
    today = date.today()

    if period == 'daily':
        start_date = today
    elif period == 'weekly':
        start_date = today - timedelta(days=7)
    else:
        start_date = today - timedelta(days=30)

    transactions = Transaction.objects.filter(
        user=request.user,
        date__gte=start_date,
        date__lte=today
    ).select_related('category')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="report_{period}.csv"'

    writer = csv.writer(response)
    writer.writerow(['Date', 'Type', 'Category', 'Amount'])

    for t in transactions:
        writer.writerow([t.date, t.transaction_type.capitalize(), t.category.name, t.amount])

    return response


@login_required(login_url='/login/')
def export_excel(request):
    period = request.GET.get('period', 'monthly')
    today = date.today()

    if period == 'daily':
        start_date = today
    elif period == 'weekly':
        start_date = today - timedelta(days=7)
    else:
        start_date = today - timedelta(days=30)

    transactions = Transaction.objects.filter(
        user=request.user,
        date__gte=start_date,
        date__lte=today
    ).select_related('category')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{period.capitalize()} Report"

    headers = ['Date', 'Type', 'Category', 'Amount']
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_num, t in enumerate(transactions, 2):
        ws.cell(row=row_num, column=1, value=str(t.date))
        ws.cell(row=row_num, column=2, value=t.transaction_type.capitalize())
        ws.cell(row=row_num, column=3, value=t.category.name)
        ws.cell(row=row_num, column=4, value=float(t.amount))

        row_fill = PatternFill(
            start_color="E2EFDA" if t.transaction_type == 'income' else "FFDDC1",
            end_color="E2EFDA" if t.transaction_type == 'income' else "FFDDC1",
            fill_type="solid"
        )
        for col in range(1, 5):
            ws.cell(row=row_num, column=col).fill = row_fill

    transaction_list = list(transactions)
    total_income = sum(float(t.amount) for t in transaction_list if t.transaction_type == 'income')
    total_expense = sum(float(t.amount) for t in transaction_list if t.transaction_type == 'expense')

    last_row = len(transaction_list) + 3
    ws.cell(row=last_row, column=3, value="Total Income").font = Font(bold=True)
    ws.cell(row=last_row, column=4, value=total_income).font = Font(bold=True)
    ws.cell(row=last_row + 1, column=3, value="Total Expense").font = Font(bold=True)
    ws.cell(row=last_row + 1, column=4, value=total_expense).font = Font(bold=True)
    ws.cell(row=last_row + 2, column=3, value="Net Balance").font = Font(bold=True)
    ws.cell(row=last_row + 2, column=4, value=total_income - total_expense).font = Font(bold=True)

    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="report_{period}.xlsx"'
    wb.save(response)

    return response
